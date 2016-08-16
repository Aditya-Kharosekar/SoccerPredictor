from openpyxl import load_workbook
import pandas as pd
 
teams = []

#This function will print the appropriate team choices, and return a list. The first element will be the index of the home team.
#The second element will be the index of the away team
def getTeamSelections():
    indexResults = []

    wb = load_workbook("Soccer Points Database.xlsx", data_only=True)
    sheet = wb["Sheet1"]

    for num in range(1, 21):        #Reading from my excel database and putting all 20 team names into the list teams[]
        cell_num = "A"+str((num))
        teams.append(sheet[cell_num].value)

    for num in range (0, 20):       #printing the home team options
        print(str(num+1) + ". " + teams[num])
    print("Enter index: ")
    home = int(input()) - 1     #I subtract 1 because the printed index of the team is actually one more than the team's index in teams[]
                                #Example: Leicester City is shown as the 8th team but it has index 7
    print("You have chosen " + teams[home] + " as the home team")
    indexResults.append(home)
    
    print("Enter index of away team: ")
    for num in range(0, 20):        #printing the away team options
       if (num!=home):              #i.e if team hasn't already been selected as the home team
           print(str(num+1) + ". " + teams[num])
    print("Enter index: ")
    away = int(input()) - 1     #I subtract 1 for the same reason as above
    print("You have chosen " + teams[away] + " as the away team")
    indexResults.append(away)  
    return indexResults    

#Takes in list of index of home team and of away team. Reads excel database and returns a list. First element will be rating of home team.
#Second element will be rating of away team
def getTeamRatings(teamchoices):
    homeTeamIndex = teamchoices[0]
    awayTeamIndex = teamchoices[1]
    wb = load_workbook("Soccer Points Database.xlsx", data_only=True)
    sheet = wb["Sheet1"]

    home_cell_num = "E" + str(homeTeamIndex+1)      # I add 1 to get the correct team's rating. The index in teamChoices[] is 1 less than the
    away_cell_num = "E" + str(awayTeamIndex+1)      # row number in the excel sheet
    homeRating = sheet[home_cell_num].value;
    awayRating = sheet[away_cell_num].value
    return [homeRating, awayRating]

# Quantifies how good the home team is at home games and how good the away team is at away games. Takes in indeces of chosen teams.
# Returns a list. First element will be the % of available points(out of 57) that the home team has won at home.
# Second element - % of available points(out of 57) that the away team has won on its travels.
def getHomeFieldAdvantage(teamchoices):
    homeTeam = teams[teamchoices[0]]
    awayTeam = teams[teamchoices[1]]

    df = pd.read_csv("2015-16.csv")

    home_df = df.loc[df["HomeTeam"]==homeTeam]
    home_win_df = home_df.loc[home_df["FTR"]=="H"]
    home_draw_df = home_df.loc[home_df["FTR"]=="D"]
    home_pts_gained = (len(home_win_df.index)*3) + (len(home_draw_df.index)*1)
    home_pointsPerc = (home_pts_gained/57)*100 #(Points gained/Available points) *100
    
    away_df = df.loc[df["AwayTeam"] ==awayTeam]
    away_win_df = away_df.loc[away_df["FTR"]=="A"]
    away_draw_df = away_df.loc[away_df["FTR"]=="D"]
    away_pts_gained = (len(away_win_df.index)*3) + (len(away_draw_df.index)*1)
    away_pointsPerc = (away_pts_gained/57)*100

    return [home_pointsPerc, away_pointsPerc]

# Takes in the list returned by getTeamRatings(), and returns win percentages of both teams in a list. First element - home team win%.
# Second element - away team win%
def getWinPercentages(teamratings, teamchoices):
    homeScore = teamratings[0]
    awayScore = teamratings[1]

    homeFieldAdvantage = getHomeFieldAdvantage(teamchoices)     #Basically tries to quantifies home field advantage

    homeRating = (0.85*homeScore)+(0.15*homeFieldAdvantage[0])
    awayRating = (0.85*awayScore)+(0.15*homeFieldAdvantage[1])
    
    x = 100/(homeRating + awayRating)
    homeWinPercentage = homeRating * x;
    awayWinPercentage = awayRating * x;

    return [homeWinPercentage, awayWinPercentage]

def main():
    print("Enter index of home team: ")
    teamChoices = getTeamSelections()
    teamRatings = getTeamRatings(teamChoices);
    percentages = getWinPercentages(teamRatings, teamChoices)
    print("Soccer Matchup Analysis")
    print(str(teams[teamChoices[0]]) + " has a " + str(round(percentages[0], 2)) + "% chance of winning");
    print(str(teams[teamChoices[1]]) + " has a " + str(round(percentages[1], 2)) + "% chance of winning");

if __name__ == '__main__':
    main()